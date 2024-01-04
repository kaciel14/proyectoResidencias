from openpyxl import load_workbook


rutaIn = ' ./../archivos/book.xlsx'
rutaOut = ' ./../archivos/resultado.xlsx'

wb = load_workbook(rutaIn)

ws = wb.active

print(ws.title)

#ws['A4'] = 'hola mundo'

for row in ws.iter_rows():
    for cell in row:
        if cell.value and '[NOMBRE]' in str(cell.value):
            cell.value = str(cell.value).replace('[NOMBRE]', 'KACIEL')

        if cell.value and '[NUMERO]' in str(cell.value):
            cell.value = 11

wb.save(rutaOut)
